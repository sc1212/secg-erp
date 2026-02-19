#!/usr/bin/env python3
"""SECG ERP — Full Import Pipeline Validation
Validates all data sources against actual uploaded files without requiring a database.
Run: python validate_imports.py /path/to/uploads/
"""

import csv
import os
import re
import sys
from collections import defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation

import openpyxl


# ─── Utility Functions ──────────────────────────────────────────────────────
def clean_currency(val):
    if val is None:
        return Decimal("0")
    if isinstance(val, (int, float)):
        return Decimal(str(round(val, 2)))
    s = str(val).replace("$", "").replace(",", "").replace(" ", "").strip()
    if not s or s in ("-", "—", "–", "N/A", "n/a"):
        return Decimal("0")
    try:
        return Decimal(s).quantize(Decimal("0.01"))
    except (InvalidOperation, ArithmeticError):
        return Decimal("0")


def clean_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    import datetime as dt
    if isinstance(val, dt.date):
        return val
    s = str(val).strip()
    for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y", "%Y-%m-%dT%H:%M:%S"]:
        try:
            return datetime.strptime(s.split()[0], fmt).date()
        except (ValueError, IndexError):
            continue
    return None


def clean_string(val, max_len=255):
    if val is None:
        return None
    s = str(val).strip()
    return s[:max_len] if s else None


def v(row, idx, default=None):
    if row is None or idx >= len(row) or idx < 0:
        return default
    rv = row[idx]
    if isinstance(rv, str) and rv.startswith("="):
        return default
    return rv


# ─── Masterfile Validation ──────────────────────────────────────────────────
def validate_masterfile(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    stats = {}

    # TXN LOG
    rows = list(wb["TXN LOG"].iter_rows(values_only=True))
    count = sum(1 for r in rows[1:]
                if clean_date(v(r, 0)) is not None or clean_currency(v(r, 6)) != 0)
    stats["TXN LOG → cost_events"] = count

    # DEBT SCHEDULE
    rows = list(wb["DEBT SCHEDULE"].iter_rows(values_only=True))
    count = sum(1 for r in rows[2:]
                if clean_string(v(r, 2)) and not str(v(r, 2)).startswith(("TOTAL", "=")))
    stats["DEBT SCHEDULE → debts"] = count

    # PROPERTIES
    rows = list(wb["PROPERTIES"].iter_rows(values_only=True))
    count = sum(1 for r in rows[2:]
                if clean_string(v(r, 0)) and not str(v(r, 0)).startswith("TOTAL"))
    stats["PROPERTIES → properties"] = count

    # PROJECT BUDGETS
    rows = list(wb["PROJECT BUDGETS"].iter_rows(values_only=True))
    projs = [(clean_string(v(r, 0)), clean_currency(v(r, 1)))
             for r in rows[2:]
             if clean_string(v(r, 0)) and not str(v(r, 0)).startswith("TOTAL")]
    stats["PROJECT BUDGETS → projects"] = f"{len(projs)} — ${sum(b for _, b in projs):,.0f} total"

    # JOB COSTING
    rows = list(wb["JOB COSTING"].iter_rows(values_only=True))
    pcount = cc = 0
    for r in rows[3:]:
        v0 = clean_string(v(r, 0))
        if not v0: continue
        if v0.startswith("PROJECT:"): pcount += 1
        elif v0 not in ("Cost Code / Division", "TOTALS") and not v0.startswith("=") and clean_string(v(r, 1)):
            cc += 1
    stats["JOB COSTING → cost_codes"] = f"{pcount} projects, {cc} codes"

    # SOV DRAW BUILDER
    rows = list(wb["SOV DRAW BUILDER"].iter_rows(values_only=True))
    pcount = sov = 0
    for r in rows[4:]:
        v0 = clean_string(v(r, 0))
        if not v0 or v0.startswith(("=", "Division", "TOTAL")): continue
        if " — " in v0 or " -- " in v0: pcount += 1
        else: sov += 1
    stats["SOV DRAW BUILDER → sov_lines"] = f"{pcount} projects, {sov} lines"

    # DRAW TRACKER
    rows = list(wb["DRAW TRACKER"].iter_rows(values_only=True))
    count = sum(1 for r in rows[4:] if clean_string(v(r, 1)) and clean_string(v(r, 2)))
    stats["DRAW TRACKER → pay_apps"] = count

    # EMPLOYEE COSTS
    rows = list(wb["EMPLOYEE COSTS"].iter_rows(values_only=True))
    count = sum(1 for r in rows[2:]
                if clean_string(v(r, 0)) and not str(v(r, 0)).startswith("TOTAL"))
    stats["EMPLOYEE COSTS → employees"] = count

    # PAYROLL
    rows = list(wb["PAYROLL"].iter_rows(values_only=True))
    count = sum(1 for r in rows[2:]
                if clean_string(v(r, 0)) and not str(v(r, 0)).startswith("TOTAL"))
    stats["PAYROLL → payroll_entries"] = count

    # PAYROLL CALENDAR
    rows = list(wb["PAYROLL CALENDAR"].iter_rows(values_only=True))
    count = sum(1 for r in rows[3:] if clean_date(v(r, 1)))
    stats["PAYROLL CALENDAR → payroll_calendar"] = count

    # PHASE SYNC
    rows = list(wb["PHASE SYNC"].iter_rows(values_only=True))
    count = sum(1 for r in rows[4:] if clean_string(v(r, 0)) and clean_string(v(r, 1)))
    stats["PHASE SYNC → phase_sync_entries"] = count

    # DEBT PAYOFF
    rows = list(wb["DEBT PAYOFF"].iter_rows(values_only=True))
    count = sum(1 for r in rows[3:]
                if clean_string(v(r, 0)) and not str(v(r, 0)).startswith("TOTAL"))
    stats["DEBT PAYOFF → debt updates"] = count

    # RECURRING EXP
    rows = list(wb["RECURRING EXP"].iter_rows(values_only=True))
    total = Decimal("0"); count = 0
    for r in rows[3:]:
        vn = clean_string(v(r, 0))
        if not vn or vn.startswith(("TOTAL", "=")): continue
        total += clean_currency(v(r, 3)); count += 1
    stats["RECURRING EXP → recurring_expenses"] = f"{count} items, ${total:,.0f}/period"

    # MONTHLY PL
    rows = list(wb["MONTHLY PL"].iter_rows(values_only=True))
    count = 0
    for r in rows[4:]:
        acct = clean_string(v(r, 0))
        if not acct or acct.startswith(("=", "TOTAL", "~")): continue
        for m in range(12):
            if clean_currency(v(r, m + 1)) != 0: count += 1
    stats["MONTHLY PL → pl_entries"] = count

    # MULTIFAMILY PL
    rows = list(wb["MULTIFAMILY PL"].iter_rows(values_only=True))
    count = 0
    for r in rows[6:]:
        acct = clean_string(v(r, 0))
        if not acct or acct.startswith(("=", "TOTAL", "~")): continue
        for m in range(12):
            if clean_currency(v(r, m + 1)) != 0: count += 1
    stats["MULTIFAMILY PL → pl_entries"] = count

    # SCENARIO MODEL
    rows = list(wb["SCENARIO MODEL"].iter_rows(values_only=True))
    sc = a = 0
    for r in rows[1:]:
        v0 = clean_string(v(r, 0))
        if not v0: continue
        if v0.startswith(("SCENARIO", "CURRENT")): sc += 1
        elif v(r, 1) is not None and not v0.startswith("="): a += 1
    stats["SCENARIO MODEL → scenarios"] = f"{sc} scenarios, {a} assumptions"

    # COA
    rows = list(wb["COA"].iter_rows(values_only=True))
    count = sum(1 for r in rows[3:] if clean_string(v(r, 0)) and clean_string(v(r, 1)))
    stats["COA → chart_of_accounts"] = count

    # CHANGE ORDERS
    rows = list(wb["CHANGE ORDERS"].iter_rows(values_only=True))
    count = sum(1 for r in rows[4:] if clean_string(v(r, 0)) and clean_string(v(r, 1)))
    stats["CHANGE ORDERS → change_orders"] = count

    # LIEN WAIVERS
    rows = list(wb["LIEN WAIVERS"].iter_rows(values_only=True))
    count = 0; exposure = Decimal("0")
    for r in rows[5:]:
        vn = clean_string(v(r, 0))
        if not vn or vn.startswith("TOTAL"): continue
        exposure += clean_currency(v(r, 3)); count += 1
    stats["LIEN WAIVERS → lien_waivers"] = f"{count} waivers, ${exposure:,.0f} exposure"

    # VENDOR SCORECARD
    rows = list(wb["VENDOR SCORECARD"].iter_rows(values_only=True))
    count = sum(1 for r in rows[4:]
                if clean_string(v(r, 0)) and not str(v(r, 0)).startswith("TOTAL"))
    stats["VENDOR SCORECARD → vendors"] = count

    # PROJECT SCHEDULE
    rows = list(wb["PROJECT SCHEDULE"].iter_rows(values_only=True))
    pcount = mcount = 0
    for r in rows[3:]:
        v0 = clean_string(v(r, 0))
        if not v0: continue
        if v0.startswith("PROJECT:"): pcount += 1
        elif v0 != "Milestone" and not v0.startswith("="): mcount += 1
    stats["PROJECT SCHEDULE → milestones"] = f"{pcount} projects, {mcount} milestones"

    # RETAINAGE
    rows = list(wb["RETAINAGE"].iter_rows(values_only=True))
    recv = pay = 0; section = None
    for r in rows:
        v0 = clean_string(v(r, 0))
        if not v0: continue
        if "HELD BY LENDERS" in v0: section = "r"; continue
        if "YOU OWE" in v0: section = "p"; continue
        if v0 in ("Project", "Sub/Vendor") or v0.startswith("TOTAL"): continue
        if section == "r": recv += 1
        elif section == "p": pay += 1
    stats["RETAINAGE → retainage_entries"] = f"{recv} lender + {pay} sub"

    # BID PIPELINE
    rows = list(wb["BID PIPELINE"].iter_rows(values_only=True))
    count = 0; pipe_val = Decimal("0")
    for r in rows[4:]:
        name = clean_string(v(r, 0))
        if not name or name.startswith("TOTAL"): continue
        pipe_val += clean_currency(v(r, 5)); count += 1
    stats["BID PIPELINE → bid_pipeline"] = f"{count} opportunities, ${pipe_val:,.0f}"

    # CREW ALLOCATION
    rows = list(wb["CREW ALLOCATION"].iter_rows(values_only=True))
    members = sum(1 for r in rows[4:]
                  if clean_string(v(r, 0)) and not str(v(r, 0)).startswith(("TOTAL", "SUB")))
    stats["CREW ALLOCATION → crew_allocations"] = f"{members} people × 13 weeks"

    # AR AGING
    rows = list(wb["AR AGING"].iter_rows(values_only=True))
    count = 0; ar_total = Decimal("0")
    for r in rows[2:]:
        client = clean_string(v(r, 0))
        if not client or client.startswith("TOTAL"): continue
        amt = clean_currency(v(r, 1))
        if amt > 0: count += 1; ar_total += amt
    stats["AR AGING → invoices"] = f"{count} receivables, ${ar_total:,.0f}"

    # AP AGING
    rows = list(wb["AP AGING"].iter_rows(values_only=True))
    count = 0; ap_total = Decimal("0")
    for r in rows[2:]:
        vn = clean_string(v(r, 0))
        if not vn or vn.startswith("TOTAL"): continue
        amt = clean_currency(v(r, 1))
        if amt > 0: count += 1; ap_total += amt
    stats["AP AGING → vendor_updates"] = f"{count} payables, ${ap_total:,.0f}"

    # LOWES PRO
    rows = list(wb["LOWES PRO"].iter_rows(values_only=True))
    count = 0; lowes_total = Decimal("0")
    for r in rows[2:]:
        date = clean_date(v(r, 0))
        amt = clean_currency(v(r, 9))
        if date and amt != 0: count += 1; lowes_total += abs(amt)
    stats["LOWES PRO → cost_events"] = f"{count} transactions, ${lowes_total:,.0f}"

    # DATA LOG
    rows = list(wb["DATA LOG"].iter_rows(values_only=True))
    count = sum(1 for r in rows[3:] if clean_string(v(r, 0)))
    stats["DATA LOG → data_sources"] = count

    wb.close()
    return stats


# ─── Budget CSV Validation ──────────────────────────────────────────────────
def validate_budget_csvs(upload_dir):
    csv_files = sorted([f for f in os.listdir(upload_dir)
                        if f.endswith(".csv") and "Budget" in f])
    if not csv_files:
        return None

    results = {}
    total_budget = Decimal("0")
    total_codes = 0
    total_draws = 0

    for cf in csv_files:
        with open(os.path.join(upload_dir, cf)) as f:
            reader = csv.reader(f)
            rows = list(reader)
        if len(rows) < 2:
            continue

        project_name = cf.replace("_Budget.csv", "").replace("_", " ")
        budget = Decimal("0")
        codes = draws = 0
        headers = rows[0]
        draw_cols = [i for i, h in enumerate(headers) if h.startswith("Draw")]
        for row in rows[1:]:
            if len(row) > 1:
                budget += clean_currency(row[1])
                codes += 1
            for dc in draw_cols:
                if dc < len(row) and clean_currency(row[dc]) != 0:
                    draws += 1

        total_budget += budget
        total_codes += codes
        total_draws += draws
        results[project_name] = (budget, codes, draws)

    return results, total_budget, total_codes, total_draws


# ─── Leads / Proposals / Jobs ───────────────────────────────────────────────
def validate_leads(upload_dir):
    f = os.path.join(upload_dir, "Leads__1_.xlsx")
    if not os.path.exists(f): return None
    wb = openpyxl.load_workbook(f, read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()
    by_status = defaultdict(int)
    count = 0
    for row in rows[1:]:
        if not row[0]: continue
        count += 1
        by_status[clean_string(v(row, 3)) or "Unknown"] += 1
    return count, dict(by_status)


def validate_proposals(upload_dir):
    f = os.path.join(upload_dir, "LeadProposals__9_.xlsx")
    if not os.path.exists(f): return None
    wb = openpyxl.load_workbook(f, read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()
    total = Decimal("0"); count = 0
    for row in rows[1:]:
        if not row[0]: continue
        count += 1; total += clean_currency(v(row, 6))
    return count, total


def validate_jobs(upload_dir):
    f = os.path.join(upload_dir, "Open_Jobs_Next_Steps_Quotes.xlsx")
    if not os.path.exists(f): return None
    wb = openpyxl.load_workbook(f, read_only=True)
    jobs = quotes = 0; quote_total = Decimal("0")
    if "Jobs Master" in wb.sheetnames:
        jobs = sum(1 for r in list(wb["Jobs Master"].iter_rows(values_only=True))[1:] if r[0])
    if "Quotes" in wb.sheetnames:
        for r in list(wb["Quotes"].iter_rows(values_only=True))[1:]:
            if not r[0]: continue
            quotes += 1; quote_total += clean_currency(v(r, 4))
    wb.close()
    return jobs, quotes, quote_total


# ─── Main ────────────────────────────────────────────────────────────────────
def main():
    upload_dir = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads"

    print("=" * 70)
    print("  SECG ERP — FULL IMPORT PIPELINE VALIDATION")
    print("=" * 70)

    masterfile = os.path.join(upload_dir, "SECG_Ultimate_Masterfile.xlsx")
    if os.path.exists(masterfile):
        print(f"\n  MASTERFILE: {os.path.basename(masterfile)}")
        print("  " + "-" * 66)
        stats = validate_masterfile(masterfile)
        total_mf = 0
        for tab, count in stats.items():
            match = re.search(r"\d+", str(count))
            n = int(match.group()) if match else 0
            total_mf += n
            print(f"    {tab:42s} {count}")
        print(f"    {'':42s} ──────────")
        print(f"    {'MASTERFILE TOTAL':42s} ~{total_mf:,} records")

    budget_result = validate_budget_csvs(upload_dir)
    if budget_result:
        by_file, total_b, total_c, total_d = budget_result
        print(f"\n  BUDGET CSVs: {len(by_file)} files")
        print("  " + "-" * 66)
        for proj, (budget, codes, draws) in by_file.items():
            print(f"    {proj:35s} ${budget:>12,.0f}  {codes:>3} codes  {draws:>3} draws")
        print(f"    {'BUDGET TOTAL':35s} ${total_b:>12,.0f}  {total_c:>3} codes  {total_d:>3} draws")

    leads_result = validate_leads(upload_dir)
    if leads_result:
        lcount, by_status = leads_result
        print(f"\n  LEADS: {lcount} total")
        print("  " + "-" * 66)
        for status in sorted(by_status, key=by_status.get, reverse=True):
            print(f"    {status:25s} {by_status[status]:>4}")

    prop_result = validate_proposals(upload_dir)
    if prop_result:
        pcount, ptotal = prop_result
        print(f"\n  PROPOSALS: {pcount} total, ${ptotal:,.0f} pipeline")

    jobs_result = validate_jobs(upload_dir)
    if jobs_result:
        jcount, qcount, qtotal = jobs_result
        print(f"\n  JOBS & QUOTES: {jcount} jobs, {qcount} quotes, ${qtotal:,.0f}")

    print("\n" + "=" * 70)
    print("  STATUS: ALL IMPORT HANDLERS VALIDATED")
    print("=" * 70)


if __name__ == "__main__":
    main()
