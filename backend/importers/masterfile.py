"""Importer for the SECG Ultimate Masterfile (38-tab Excel workbook).

Column mappings verified against actual SECG_Ultimate_Masterfile.xlsx
uploaded 2026-02-19.
"""

from datetime import datetime
from decimal import Decimal
from typing import Optional
import re

import openpyxl
from sqlalchemy.orm import Session

from backend.importers.base import BaseImporter, ImportResult
from backend.models.core import (
    ChangeOrder, ChangeOrderStatus, CostCode, CostEvent, CostEventSource,
    CostEventType, Employee, Invoice, InvoiceStatus, Payment, PayApp,
    PayAppLine, Project, ProjectStatus, ProjectType, Quote, SOVLine, Vendor,
)
from backend.models.extended import (
    BidPipeline, BidStatus, CashForecastLine, CashSnapshot, ChartOfAccounts,
    CrewAllocation, DataSource, Debt, DebtType, Lead, LienWaiver,
    MilestoneStatus, PLEntry, PayrollCalendar, PayrollEntry, PhaseSyncEntry,
    ProjectMilestone, Property, RecurringExpense, RecurringFrequency,
    RetainageEntry, Scenario, ScenarioAssumption,
)

SKIP_TABS = {
    "TAB INDEX", "KPI", "DASHBOARD", "WIP SCHEDULE",
    "PAYMENT WATERFALL", "WEEKLY DIGEST", "CASH POSITION",
}


class MasterfileImporter(BaseImporter):
    source_name = "masterfile"
    source_type = "masterfile_import"

    def __init__(self, session: Session, file_path: str, batch_id: Optional[str] = None):
        super().__init__(session, batch_id)
        self.file_path = file_path

    def run(self) -> ImportResult:
        wb = openpyxl.load_workbook(self.file_path, read_only=True)
        tab_handlers = {
            "TXN LOG": self._import_txn_log,
            "DEBT SCHEDULE": self._import_debts,
            "PROPERTIES": self._import_properties,
            "PROJECT BUDGETS": self._import_project_budgets,
            "JOB COSTING": self._import_job_costing,
            "SOV DRAW BUILDER": self._import_sov,
            "DRAW TRACKER": self._import_draw_tracker,
            "EMPLOYEE COSTS": self._import_employee_costs,
            "PAYROLL": self._import_payroll,
            "PAYROLL CALENDAR": self._import_payroll_calendar,
            "DAILY INPUTS": self._import_daily_inputs,
            "CASH FLOW 13WK": self._import_cash_forecast,
            "PHASE SYNC": self._import_phase_sync,
            "DEBT PAYOFF": self._import_debt_payoff,
            "RECURRING EXP": self._import_recurring_expenses,
            "MONTHLY PL": self._import_monthly_pl,
            "MULTIFAMILY PL": self._import_multifamily_pl,
            "SCENARIO MODEL": self._import_scenarios,
            "COA": self._import_chart_of_accounts,
            "DATA LOG": self._import_data_log,
            "CHANGE ORDERS": self._import_change_orders,
            "LIEN WAIVERS": self._import_lien_waivers,
            "VENDOR SCORECARD": self._import_vendor_scorecard,
            "PROJECT SCHEDULE": self._import_project_schedule,
            "RETAINAGE": self._import_retainage,
            "BID PIPELINE": self._import_bid_pipeline,
            "CREW ALLOCATION": self._import_crew_allocation,
            "AR AGING": self._import_ar_aging,
            "AP AGING": self._import_ap_aging,
            "LOWES PRO": self._import_lowes_pro,
            "VICTORY CROSSINGS": self._import_victory_crossings,
        }
        for tab_name in wb.sheetnames:
            if tab_name in SKIP_TABS:
                print(f"  skip {tab_name} (computed)")
                continue
            handler = tab_handlers.get(tab_name)
            if handler:
                print(f"  importing {tab_name}...")
                try:
                    ws = wb[tab_name]
                    rows = list(ws.iter_rows(values_only=True))
                    handler(rows)
                    self.session.commit()
                    print(f"    ok {tab_name}")
                except Exception as e:
                    self.result.errors.append(f"Tab '{tab_name}': {e}")
                    self.session.rollback()
                    print(f"    FAIL {tab_name}: {e}")
            else:
                print(f"  no handler: {tab_name}")
        wb.close()
        self.log_data_source(self.result.total_processed)
        self.session.commit()
        self.result.finish()
        return self.result

    # helpers
    def _val(self, row, idx, default=None):
        if row is None or idx >= len(row) or idx < 0:
            return default
        v = row[idx]
        if isinstance(v, str) and v.startswith("="):
            return default
        return v

    def _is_section_header(self, val):
        if val is None: return False
        s = str(val).strip()
        return s.startswith("=") or s.startswith("TOTAL") or s.startswith("~")

    @staticmethod
    def _extract_project_code(name):
        if not name: return "UNKNOWN"
        paren = re.search(r'\((\w{2,6})\)', name)
        if paren: return paren.group(1)
        if " -- " in name or " — " in name:
            sep = " -- " if " -- " in name else " — "
            code = name.split(sep)[0].strip()
            if len(code) <= 8: return code.replace(" ","")
        name_lower = name.lower()
        code_map = {
            "walnut grove lot 1": "WG1", "walnut grove lot 2": "WG2",
            "walnut grove lot 3": "WG3", "walnut grove 1": "WG1",
            "walnut grove 2": "WG2", "walnut grove 3": "WG3",
            "wg1": "WG1", "wg2": "WG2", "wg3": "WG3",
            "205 kerr": "KA2", "203 kerr": "KA1",
            "ka1": "KA1", "ka2": "KA2",
            "rockvale": "RV1", "rv1": "RV1",
            "veterans": "VET1", "vet1": "VET1",
            "2145 cason": "CS1", "2147 cason": "CS2",
            "cs1": "CS1", "cs2": "CS2",
            "miller": "MILLER", "legens": "LEG1",
        }
        for pattern, code in code_map.items():
            if pattern in name_lower: return code
        return name.split()[0][:8].upper()

    # === TXN LOG -> cost_events (3336 rows) ===
    # Row 0 header: Date[0] Source[1] Type[2] Account[3] Vendor/Payee[4]
    #   Description/Memo[5] Amount[6] Division[7] Project[8]
    #   QBO Category[9] QBO Acct Code[10] Payment Method[11]
    #   User/Employee[12] Buyer[13] PO Number[14] Is MF?[15]
    def _import_txn_log(self, rows):
        if len(rows) < 2: return
        for row in rows[1:]:
            date = self.clean_date(self._val(row, 0))
            amount = self.clean_currency(self._val(row, 6))
            if date is None and amount == 0:
                self.result.skipped += 1
                continue
            project_name = self.clean_string(self._val(row, 8))
            project_id = self.get_or_create_project(project_name[:20], project_name) if project_name else None
            vendor_name = self.clean_string(self._val(row, 4))
            vendor_id = self.get_or_create_vendor(vendor_name) if vendor_name else None
            src_raw = self.clean_string(self._val(row, 1)) or ""
            src_map = {"ramp": CostEventSource.ramp_import, "qbo": CostEventSource.qbo_sync,
                       "lowes": CostEventSource.lowes_import, "home depot": CostEventSource.homedepot_import,
                       "buildertrend": CostEventSource.buildertrend_import}
            source = CostEventSource.masterfile_import
            for k, v in src_map.items():
                if k in src_raw.lower(): source = v; break
            event = CostEvent(
                project_id=project_id, vendor_id=vendor_id, date=date, amount=amount,
                description=self.clean_string(self._val(row, 5), 500),
                reference_number=self.clean_string(self._val(row, 10)),
                po_number=self.clean_string(self._val(row, 14)),
                source=source, source_ref=src_raw, import_batch=self.batch_id,
                notes=self.clean_string(self._val(row, 9)),
            )
            self.session.add(event)
            self.result.created += 1
            if self.result.created % 500 == 0: self.session.flush()

    # === DEBT SCHEDULE -> debts (59 rows) ===
    # Row 1 header but data shifted: [0]=Category [1]=Priority [2]=Creditor
    #   [3]=Amount [4]=Detail [5]=Risk [6]=TOTAL(formula) [7]=Status [8]=Notes [9]=SOURCE
    def _import_debts(self, rows):
        if len(rows) < 3: return
        for row in rows[2:]:
            creditor = self.clean_string(self._val(row, 2))
            if not creditor: self.result.skipped += 1; continue
            if creditor.startswith("TOTAL") or creditor.startswith("="): self.result.skipped += 1; continue
            existing = self.session.query(Debt).filter(Debt.name == creditor).first()
            if existing: self.result.skipped += 1; continue
            category = self.clean_string(self._val(row, 0))
            dt = DebtType.other
            if category:
                cl = category.lower()
                if "credit card" in cl: dt = DebtType.credit_card
                elif "construction" in cl or "loan" in cl: dt = DebtType.construction_loan
            amount = self.clean_currency(self._val(row, 3))
            status = self.clean_string(self._val(row, 7))
            debt = Debt(
                name=creditor, lender=category, debt_type=dt,
                current_balance=amount, original_balance=amount,
                is_active=status not in ("SETTLED", "PAID") if status else True,
                notes="; ".join(filter(None, [
                    self.clean_string(self._val(row, 4)),
                    self.clean_string(self._val(row, 5)),
                    status, self.clean_string(self._val(row, 8)),
                ])),
            )
            self.session.add(debt)
            self.result.created += 1

    # === PROPERTIES -> properties (16 rows) ===
    # Row 1 header: Property[0] SqFt[1] Construction Budget[2] Current Debt[3]
    #   Remaining Draw[4] ARV[5] Equity[6](f) $/SqFt[7](f) LTV[8](f)
    #   Exit Strategy[9] Lender[10] Status[11]
    def _import_properties(self, rows):
        if len(rows) < 3: return
        for row in rows[2:]:
            name = self.clean_string(self._val(row, 0))
            if not name or name.startswith("TOTAL"): self.result.skipped += 1; continue
            existing = self.session.query(Property).filter(Property.address == name).first()
            if existing: self.result.skipped += 1; continue
            budget = self.clean_currency(self._val(row, 2))
            debt_val = self.clean_currency(self._val(row, 3))
            arv = self.clean_currency(self._val(row, 5))
            prop = Property(
                address=name, state="TN",
                purchase_price=budget, current_value=arv, arv=arv,
                equity=arv - debt_val if arv > 0 and debt_val > 0 else None,
                ltv=Decimal(str(round(float(debt_val)/float(arv),4))) if arv > 0 else None,
                exit_strategy=self.clean_string(self._val(row, 9)),
                notes=f"Lender: {self._val(row, 10) or ''} | SqFt: {self._val(row, 1) or ''} | Status: {self._val(row, 11) or ''}",
            )
            self.session.add(prop)
            self.result.created += 1

    # === PROJECT BUDGETS -> projects (13 rows) ===
    # Row 1 header: Project[0] Total Budget[1] Draws Released[2]
    #   Remaining Balance[3](f) % Complete[4](f) Draws#[5] PM[6] Status/Notes[7]
    def _import_project_budgets(self, rows):
        if len(rows) < 3: return
        for row in rows[2:]:
            name = self.clean_string(self._val(row, 0))
            if not name or name.startswith("TOTAL"): self.result.skipped += 1; continue
            budget = self.clean_currency(self._val(row, 1))
            released = self.clean_currency(self._val(row, 2))
            pm = self.clean_string(self._val(row, 6))
            code = self._extract_project_code(name)
            project = self.session.query(Project).filter(Project.code == code).first()
            if project:
                project.budget_total = budget
                project.contract_amount = released
                project.project_manager = pm
                project.notes = self.clean_string(self._val(row, 7))
                self.result.updated += 1
            else:
                project = Project(code=code, name=name, status=ProjectStatus.active,
                    budget_total=budget, contract_amount=released,
                    project_manager=pm, state="TN",
                    notes=self.clean_string(self._val(row, 7)))
                self.session.add(project)
                self.result.created += 1
            self.session.flush()

    # === JOB COSTING -> cost_codes (135 rows, multi-project sections) ===
    # Section start: "PROJECT: WG1" then header: Cost Code/Division[0] Code[1]
    #   Budget[2] Actual Cost[3] Variance($)[4](f) Variance(%)[5](f)
    #   Committed[6] Open POs[7] Est to Complete[8](f) Est at Complete[9](f)
    def _import_job_costing(self, rows):
        if len(rows) < 5: return
        current_project_id = None
        for row in rows[3:]:
            val0 = self.clean_string(self._val(row, 0))
            if not val0: continue
            if val0.startswith("PROJECT:"):
                code = val0.replace("PROJECT:", "").strip().split()[0]
                p = self.session.query(Project).filter(Project.code == code).first()
                current_project_id = p.id if p else self.get_or_create_project(code, code)
                continue
            if val0 in ("Cost Code / Division", "TOTALS") or val0.startswith("="): continue
            if current_project_id is None: continue
            code_str = self.clean_string(self._val(row, 1))
            if not code_str: continue
            budget = self.clean_currency(self._val(row, 2))
            actual = self.clean_currency(self._val(row, 3))
            committed = self.clean_currency(self._val(row, 6))
            existing = self.session.query(CostCode).filter(
                CostCode.project_id == current_project_id, CostCode.code == code_str).first()
            if existing:
                existing.description = val0; existing.budget_amount = budget
                existing.actual_amount = actual; existing.committed_amount = committed
                existing.variance = budget - actual if budget > 0 else Decimal("0")
                self.result.updated += 1
            else:
                cc = CostCode(project_id=current_project_id, code=code_str,
                    description=val0, budget_amount=budget, actual_amount=actual,
                    committed_amount=committed, category=val0,
                    variance=budget - actual if budget > 0 else Decimal("0"))
                self.session.add(cc)
                self.result.created += 1

    # === SOV DRAW BUILDER -> sov_lines (219 rows, multi-project) ===
    # Row 3 header: Division[0] Scheduled Value[1] Prior Draws[2] Balance[3]
    #   % Complete[4] # of Draws[5] Work This Period[6] Total Completed[7](f)
    #   Stored Materials[8] Total+Stored[9](f) % of Total[10](f)
    #   Retainage(10%)[11](f) Net Draw This Period[12](f) SOURCE[13]
    def _import_sov(self, rows):
        if len(rows) < 5: return
        current_project_id = None; line_num = 0
        for row in rows[4:]:
            val0 = self.clean_string(self._val(row, 0))
            if not val0: continue
            if val0.startswith("=") or val0 == "Division": continue
            # detect project header
            if " — " in val0 or " -- " in val0:
                code = self._extract_project_code(val0)
                p = self.session.query(Project).filter(Project.code == code).first()
                current_project_id = p.id if p else self.get_or_create_project(code, val0)
                line_num = 0; continue
            if val0.startswith("TOTAL"): continue
            if current_project_id is None: continue
            line_num += 1
            sched = self.clean_currency(self._val(row, 1))
            pct_raw = self._val(row, 4)
            pct = Decimal("0")
            if pct_raw is not None and not isinstance(pct_raw, str):
                try: pct = Decimal(str(round(float(pct_raw)*100, 2)))
                except: pass
            sov = SOVLine(project_id=current_project_id, line_number=line_num,
                description=val0, scheduled_value=sched,
                previous_billed=self.clean_currency(self._val(row, 2)),
                current_billed=self.clean_currency(self._val(row, 6)),
                balance_to_finish=self.clean_currency(self._val(row, 3)),
                stored_materials=self.clean_currency(self._val(row, 8)),
                percent_complete=pct)
            self.session.add(sov)
            self.result.created += 1

    # === DRAW TRACKER -> pay_apps (77 rows) ===
    # Data shifted: Seq[0] Project[1] Phase[2] Amount[3] Status[4] ...
    def _import_draw_tracker(self, rows):
        if len(rows) < 5: return
        for row in rows[4:]:
            project_code = self.clean_string(self._val(row, 1))
            phase = self.clean_string(self._val(row, 2))
            amount = self.clean_currency(self._val(row, 3))
            status = self.clean_string(self._val(row, 4))
            if not project_code or not phase: self.result.skipped += 1; continue
            p = self.session.query(Project).filter(Project.code == project_code).first()
            pid = p.id if p else self.get_or_create_project(project_code, project_code)
            draw_num = 0
            if phase and "draw" in phase.lower():
                for part in phase.split():
                    try: draw_num = int(part); break
                    except: continue
            existing = self.session.query(PayApp).filter(
                PayApp.project_id == pid, PayApp.pay_app_number == draw_num).first()
            if existing: self.result.skipped += 1; continue
            pa_status = "paid" if status and "RELEASED" in status.upper() else "draft"
            pa = PayApp(project_id=pid, pay_app_number=draw_num or self.clean_int(self._val(row, 0)),
                amount_requested=amount, amount_approved=amount if pa_status == "paid" else Decimal("0"),
                net_payment=amount if pa_status == "paid" else Decimal("0"),
                status=pa_status, notes=phase)
            self.session.add(pa)
            self.result.created += 1

    # === EMPLOYEE COSTS -> employees (14 rows) ===
    # Row 1 header (data shifted): Employee[0] Title[1] Salary[2]
    #   Employer Taxes[3] Vehicle[4] ...
    def _import_employee_costs(self, rows):
        if len(rows) < 3: return
        for row in rows[2:]:
            name = self.clean_string(self._val(row, 0))
            if not name or name.startswith("TOTAL"): self.result.skipped += 1; continue
            title = self.clean_string(self._val(row, 1))
            salary = self.clean_currency(self._val(row, 2))
            parts = name.split(None, 1)
            first = parts[0]; last = parts[1] if len(parts) > 1 else ""
            existing = self.session.query(Employee).filter(
                Employee.first_name == first, Employee.last_name == last).first()
            if existing:
                existing.role = title; existing.salary = salary; self.result.updated += 1
            else:
                emp = Employee(first_name=first, last_name=last, role=title, salary=salary, is_active=True)
                self.session.add(emp); self.result.created += 1

    # === PAYROLL -> payroll_entries + employee updates (13 rows) ===
    # Row 1 header: Employee[0] Title[1] Hourly Rate[2] Weekly Gross[3]
    #   Annual Salary[4] YTD Gross[5](f) YTD Net Pay[6] YTD Emp Taxes[7]
    #   YTD Employer Taxes[8] YTD Reimb[9] Total YTD[10] Status[11]
    def _import_payroll(self, rows):
        if len(rows) < 3: return
        for row in rows[2:]:
            name = self.clean_string(self._val(row, 0))
            if not name or name.startswith("TOTAL"): self.result.skipped += 1; continue
            title = self.clean_string(self._val(row, 1))
            hourly = self.clean_currency(self._val(row, 2))
            annual = self.clean_currency(self._val(row, 4))
            parts = name.split(None, 1)
            first = parts[0]; last = parts[1] if len(parts) > 1 else ""
            emp = self.session.query(Employee).filter(
                Employee.first_name == first, Employee.last_name == last).first()
            if emp:
                if hourly > 0: emp.hourly_rate = hourly
                if annual > 0: emp.salary = annual
                emp.role = title or emp.role
                self.result.updated += 1
            else:
                emp = Employee(first_name=first, last_name=last, role=title,
                    hourly_rate=hourly if hourly > 0 else None,
                    salary=annual if annual > 0 else None, is_active=True)
                self.session.add(emp); self.session.flush()
                self.result.created += 1
            ytd_er_tax = self.clean_currency(self._val(row, 8))
            if annual > 0 and emp.id:
                entry = PayrollEntry(employee_id=emp.id,
                    pay_period_start=datetime(2025,1,1).date(),
                    pay_period_end=datetime(2025,12,31).date(),
                    gross_pay=annual, net_pay=self.clean_currency(self._val(row, 6)),
                    employer_taxes=ytd_er_tax, total_cost=annual + ytd_er_tax,
                    notes=self.clean_string(self._val(row, 11)))
                self.session.add(entry)

    # === PAYROLL CALENDAR -> payroll_calendar (30 rows) ===
    # Row 2 header: Pay#[0] Pay Date[1] Est Gross[2](f) Employer Taxes[3](f)
    #   Total Cash Needed[4](f) Draw Expected?[5] Cash Source/Notes[6]
    def _import_payroll_calendar(self, rows):
        if len(rows) < 4: return
        for row in rows[3:]:
            pay_date = self.clean_date(self._val(row, 1))
            if not pay_date: self.result.skipped += 1; continue
            existing = self.session.query(PayrollCalendar).filter(
                PayrollCalendar.pay_date == pay_date).first()
            if existing: self.result.skipped += 1; continue
            cal = PayrollCalendar(pay_date=pay_date, period_start=pay_date,
                period_end=pay_date, status="scheduled",
                notes=self.clean_string(self._val(row, 6)))
            self.session.add(cal); self.result.created += 1

    # === DAILY INPUTS -> cash_snapshots (71 rows, form layout) ===
    def _import_daily_inputs(self, rows):
        if len(rows) < 10: return
        snap = CashSnapshot(snapshot_date=datetime.now().date(),
            account_name="DAILY INPUTS Snapshot",
            balance=self.clean_currency(self._val(rows[7] if len(rows) > 7 else None, 1)),
            notes="Imported from DAILY INPUTS tab")
        self.session.add(snap); self.result.created += 1

    # === CASH FLOW 13WK -> cash_forecast_lines (45 rows) ===
    # Row 2 header: Category[0] Monthly Avg[1] Wk1[2]..Wk13[14] SOURCE[15]
    def _import_cash_forecast(self, rows):
        if len(rows) < 4: return
        header = rows[2]
        week_dates = []
        for i in range(2, min(15, len(header) if header else 0)):
            h = str(header[i]) if header[i] else ""
            d = None
            if "\n" in h:
                dp = h.split("\n")[-1]
                d = self.clean_date(f"{dp}/2026")
            week_dates.append((i, d))
        for row in rows[3:]:
            cat = self.clean_string(self._val(row, 0))
            if not cat or cat.startswith("=") or cat.startswith("TOTAL"): continue
            for ci, wd in week_dates:
                val = self._val(row, ci)
                amt = self.clean_currency(val)
                if amt == 0: continue
                line = CashForecastLine(week_starting=wd or datetime.now().date(),
                    category=cat, amount_in=amt if amt > 0 else Decimal("0"),
                    amount_out=abs(amt) if amt < 0 else Decimal("0"), net=amt)
                self.session.add(line); self.result.created += 1

    # === PHASE SYNC -> phase_sync_entries (25 rows) ===
    # Row 3 header: Project[0] Phase[1] Material Type[2] Supplier[3]
    #   Supplier Terms[4] PO Date[5] ... (24 cols)
    def _import_phase_sync(self, rows):
        if len(rows) < 5: return
        for row in rows[4:]:
            pc = self.clean_string(self._val(row, 0))
            phase = self.clean_string(self._val(row, 1))
            if not pc or not phase: self.result.skipped += 1; continue
            p = self.session.query(Project).filter(Project.code == pc).first()
            pid = p.id if p else self.get_or_create_project(pc, pc)
            entry = PhaseSyncEntry(project_id=pid, phase_name=phase,
                status=self.clean_string(self._val(row, 4)),
                planned_start=self.clean_date(self._val(row, 5)),
                planned_end=self.clean_date(self._val(row, 8)),
                notes=f"Material: {self._val(row, 2) or ''} | Supplier: {self._val(row, 3) or ''}")
            self.session.add(entry); self.result.created += 1

    # === DEBT PAYOFF -> updates debts (24 rows) ===
    # Row 2 header: Creditor[0] Amount Owed[1] Settlement Offer[2]
    #   Settlement%[3](f) Savings[4](f) Lien Risk[5] Priority[6](f)
    #   Payoff Order[7] Strategy Notes[8]
    def _import_debt_payoff(self, rows):
        if len(rows) < 4: return
        for row in rows[3:]:
            creditor = self.clean_string(self._val(row, 0))
            if not creditor or creditor.startswith("TOTAL"): self.result.skipped += 1; continue
            debt = self.session.query(Debt).filter(
                Debt.name.ilike(f"%{creditor.split('(')[0].strip()[:20]}%")).first()
            if debt:
                settlement = self.clean_currency(self._val(row, 2))
                strategy = self.clean_string(self._val(row, 8))
                extras = []
                if settlement > 0: extras.append(f"Settlement: ${settlement:,.0f}")
                if strategy: extras.append(f"Strategy: {strategy}")
                order = self._val(row, 7)
                if order: extras.append(f"Payoff Order: {order}")
                if extras: debt.notes = (debt.notes or "") + " | " + " | ".join(extras)
                self.result.updated += 1
            else:
                self.result.skipped += 1

    # === RECURRING EXP -> recurring_expenses (37 rows) ===
    # Row 2 header: Vendor/Expense[0] Category[1] Frequency[2] Amount[3]
    #   Annual Total[4](f) Auto-Pay?[5] Due Day[6] Payment Method[7]
    #   Status[8] Notes[9]
    def _import_recurring_expenses(self, rows):
        if len(rows) < 4: return
        for row in rows[3:]:
            vn = self.clean_string(self._val(row, 0))
            if not vn or vn.startswith("TOTAL") or vn.startswith("="): self.result.skipped += 1; continue
            amount = self.clean_currency(self._val(row, 3))
            freq_str = (self.clean_string(self._val(row, 2)) or "").lower()
            freq = None
            if "biweekly" in freq_str or "bi-weekly" in freq_str: freq = RecurringFrequency.biweekly
            elif "weekly" in freq_str: freq = RecurringFrequency.weekly
            elif "month" in freq_str: freq = RecurringFrequency.monthly
            elif "quarter" in freq_str: freq = RecurringFrequency.quarterly
            elif "annual" in freq_str: freq = RecurringFrequency.annually
            vid = self.get_or_create_vendor(vn)
            exp = RecurringExpense(vendor_id=vid, description=vn, amount=amount, frequency=freq,
                is_active=(self.clean_string(self._val(row, 8)) or "").upper() != "CANCELLED",
                notes=f"Cat: {self._val(row, 1) or ''} | AutoPay: {self._val(row, 5) or ''} | Method: {self._val(row, 7) or ''} | Status: {self._val(row, 8) or ''} | {self._val(row, 9) or ''}")
            self.session.add(exp); self.result.created += 1

    # === MONTHLY PL -> pl_entries division=company_wide (51 rows) ===
    # Row 2 header: MONTH->[0] Jan[1]..Dec[12] YTD TOTAL[13] SOURCE[14]
    def _import_monthly_pl(self, rows):
        self._import_pl_pivoted(rows, "company_wide", data_start=4)

    # === MULTIFAMILY PL -> pl_entries division=multifamily (86 rows) ===
    # Row 4 header, data from row 6
    def _import_multifamily_pl(self, rows):
        self._import_pl_pivoted(rows, "multifamily", data_start=6)

    def _import_pl_pivoted(self, rows, division, data_start):
        if len(rows) < data_start + 1: return
        for row in rows[data_start:]:
            acct = self.clean_string(self._val(row, 0))
            if not acct or acct.startswith("=") or acct.startswith("TOTAL") or acct.startswith("~"): continue
            for m_idx in range(12):
                col = m_idx + 1
                val = self._val(row, col)
                amount = self.clean_currency(val)
                if amount == 0: continue
                entry = PLEntry(period_year=2025, period_month=m_idx + 1,
                    division=division, account_name=acct, amount=amount, is_budget=False)
                self.session.add(entry); self.result.created += 1

    # === SCENARIO MODEL -> scenarios + assumptions (56 rows) ===
    def _import_scenarios(self, rows):
        if len(rows) < 12: return
        current_id = None
        for row in rows[1:]:
            val0 = self.clean_string(self._val(row, 0))
            if not val0: continue
            if val0.startswith("SCENARIO") or val0.startswith("CURRENT STATE"):
                existing = self.session.query(Scenario).filter(Scenario.name == val0).first()
                if existing: current_id = existing.id; self.result.skipped += 1
                else:
                    sc = Scenario(name=val0, is_baseline=val0.startswith("CURRENT"))
                    self.session.add(sc); self.session.flush()
                    current_id = sc.id; self.result.created += 1
                continue
            if current_id is None: continue
            var_val = self._val(row, 1)
            if var_val is not None:
                a = ScenarioAssumption(scenario_id=current_id, variable_name=val0,
                    variable_value=str(var_val))
                self.session.add(a); self.result.created += 1

    # === COA -> chart_of_accounts (98 rows) ===
    # Row 2 header: Acct#[0] Account Name[1] Type[2] Detail Type/Notes[3]
    def _import_chart_of_accounts(self, rows):
        if len(rows) < 4: return
        for row in rows[3:]:
            acct_num = self.clean_string(self._val(row, 0), 20)
            name = self.clean_string(self._val(row, 1))
            if not acct_num or not name: self.result.skipped += 1; continue
            existing = self.session.query(ChartOfAccounts).filter(
                ChartOfAccounts.account_number == acct_num).first()
            if existing: self.result.skipped += 1; continue
            coa = ChartOfAccounts(account_number=acct_num, name=name,
                account_type=self.clean_string(self._val(row, 2)),
                notes=self.clean_string(self._val(row, 3)))
            self.session.add(coa); self.result.created += 1

    # === DATA LOG -> data_sources (20 rows) ===
    # Row 2 header: File Name[0] Type[1] Used In Tab(s)[2] Key Data[3]
    #   Records[4] Date Range[5] Status[6]
    def _import_data_log(self, rows):
        if len(rows) < 4: return
        for row in rows[3:]:
            name = self.clean_string(self._val(row, 0))
            if not name: continue
            ds = DataSource(name=name, source_type=self.clean_string(self._val(row, 1)),
                status=self.clean_string(self._val(row, 6)),
                notes=f"Tabs: {self._val(row, 2) or ''} | Data: {self._val(row, 3) or ''}")
            self.session.add(ds); self.result.created += 1

    # === CHANGE ORDERS -> change_orders (47 rows) ===
    # Row 3 header: CO#[0] Project[1] Date Submitted[2] Description[3]
    #   Requested By[4] Material$[5] Labor$[6] Sub$[7] Total CO Value[8](f)
    #   Schedule Impact[9] Submitted To[10] Status[11] Approval Date[12]
    #   Approved Amount[13] Variance[14](f) Reason Code[15] Notes[16]
    def _import_change_orders(self, rows):
        if len(rows) < 5: return
        for row in rows[4:]:
            co_num = self.clean_string(self._val(row, 0), 20)
            project_code = self.clean_string(self._val(row, 1))
            if not co_num or not project_code: self.result.skipped += 1; continue
            p = self.session.query(Project).filter(Project.code == project_code).first()
            pid = p.id if p else self.get_or_create_project(project_code, project_code)
            mat = self.clean_currency(self._val(row, 5))
            labor = self.clean_currency(self._val(row, 6))
            sub = self.clean_currency(self._val(row, 7))
            total = mat + labor + sub
            status_raw = (self.clean_string(self._val(row, 11)) or "").lower()
            status = ChangeOrderStatus.draft
            if "approved" in status_raw: status = ChangeOrderStatus.approved
            elif "pending" in status_raw: status = ChangeOrderStatus.pending_approval
            elif "rejected" in status_raw: status = ChangeOrderStatus.rejected
            co = ChangeOrder(project_id=pid, co_number=co_num,
                title=self.clean_string(self._val(row, 3)),
                description=self.clean_string(self._val(row, 3), 2000),
                amount=total, status=status,
                requested_by=self.clean_string(self._val(row, 4)),
                date_submitted=self.clean_date(self._val(row, 2)),
                date_approved=self.clean_date(self._val(row, 12)),
                notes=f"Mat: ${mat:,.0f} | Lab: ${labor:,.0f} | Sub: ${sub:,.0f} | Reason: {self._val(row, 15) or ''} | {self._val(row, 16) or ''}")
            self.session.add(co); self.result.created += 1

    # === LIEN WAIVERS -> lien_waivers (53 rows) ===
    # Row 4 header: Vendor/Sub[0] Project[1] Draw#[2] Payment Amount[3]
    #   Conditional Waiver?[4] Conditional Date[5] Payment Sent Date[6]
    #   Payment Cleared?[7] Unconditional Waiver?[8] Unconditional Date[9]
    #   Waiver Covers Through[10] Remaining Owed[11] Risk Level[12] Notes[13]
    def _import_lien_waivers(self, rows):
        if len(rows) < 6: return
        for row in rows[5:]:
            vn = self.clean_string(self._val(row, 0))
            if not vn or vn.startswith("TOTAL"): self.result.skipped += 1; continue
            pc = self.clean_string(self._val(row, 1))
            vid = self.get_or_create_vendor(vn)
            p = self.session.query(Project).filter(Project.code == pc).first() if pc else None
            pid = p.id if p else None
            cond = self.clean_string(self._val(row, 4))
            uncond = self.clean_string(self._val(row, 8))
            wtype = "unconditional" if uncond else ("conditional" if cond else None)
            lw = LienWaiver(project_id=pid, vendor_id=vid, waiver_type=wtype,
                amount=self.clean_currency(self._val(row, 3)),
                through_date=self.clean_date(self._val(row, 10)),
                received_date=self.clean_date(self._val(row, 9)) or self.clean_date(self._val(row, 5)),
                notes=f"Risk: {self._val(row, 12) or ''} | Draw: {self._val(row, 2) or ''} | {self._val(row, 13) or ''}")
            self.session.add(lw); self.result.created += 1

    # === VENDOR SCORECARD -> vendors (49 rows) ===
    # Row 3 header: Vendor/Sub[0] Trade[1] Phone[2] Email[3]
    #   Insurance Expires[4] Insurance Current?[5] Quality(1-5)[6]
    #   On-Time(1-5)[7] Communication(1-5)[8] Price(1-5)[9]
    #   Lien Waiver Compliance(1-5)[10] Composite Score[11](f)
    #   Total Paid YTD[12] Total Still Owed[13] Last Payment Date[14]
    #   Projects Worked[15] Preferred?[16] Notes/Issues[17]
    def _import_vendor_scorecard(self, rows):
        if len(rows) < 5: return
        for row in rows[4:]:
            name = self.clean_string(self._val(row, 0))
            if not name or name.startswith("TOTAL"): self.result.skipped += 1; continue
            vendor = self.session.query(Vendor).filter(Vendor.name == name).first()
            if not vendor:
                vendor = Vendor(name=name)
                self.session.add(vendor); self.session.flush()
            vendor.trade = self.clean_string(self._val(row, 1))
            phone_raw = self._val(row, 2)
            if phone_raw is not None:
                ps = str(phone_raw).replace(".0", "")
                if len(ps) >= 10: vendor.phone = ps
            email_raw = self._val(row, 3)
            if email_raw is not None:
                es = str(email_raw).replace(".0", "")
                if "@" in es: vendor.email = es
                elif len(es) >= 10 and not vendor.phone: vendor.phone = es
            vendor.insurance_expiry = self.clean_date(self._val(row, 4))
            vendor.score_quality = self.clean_int(self._val(row, 6)) or None
            vendor.score_timeliness = self.clean_int(self._val(row, 7)) or None
            vendor.score_communication = self.clean_int(self._val(row, 8)) or None
            vendor.score_price = self.clean_int(self._val(row, 9)) or None
            vendor.notes = self.clean_string(self._val(row, 17), 2000)
            self.result.updated += 1

    # === PROJECT SCHEDULE -> project_milestones (134 rows, multi-project) ===
    # Section: "PROJECT: WG1 — Walnut Grove 1" then header:
    #   Milestone[0] Planned Date[1] Actual Date[2] Days Variance[3](f)
    #   Status[4](f) Responsible[5] Notes[6]
    def _import_project_schedule(self, rows):
        if len(rows) < 5: return
        current_pid = None; sort_order = 0
        for row in rows[3:]:
            val0 = self.clean_string(self._val(row, 0))
            if not val0: continue
            if val0.startswith("PROJECT:"):
                code = self._extract_project_code(val0.replace("PROJECT:", "").strip())
                p = self.session.query(Project).filter(Project.code == code).first()
                current_pid = p.id if p else self.get_or_create_project(code, val0)
                sort_order = 0; continue
            if val0 == "Milestone" or val0.startswith("="): continue
            if current_pid is None: continue
            sort_order += 1
            planned = self.clean_date(self._val(row, 1))
            actual = self.clean_date(self._val(row, 2))
            if actual: status = MilestoneStatus.completed
            elif planned and planned < datetime.now().date(): status = MilestoneStatus.delayed
            else: status = MilestoneStatus.not_started
            ms = ProjectMilestone(project_id=current_pid, task_name=val0,
                status=status, planned_start=planned, planned_end=planned,
                actual_start=actual, actual_end=actual,
                assigned_to=self.clean_string(self._val(row, 5)),
                sort_order=sort_order, notes=self.clean_string(self._val(row, 6)))
            self.session.add(ms); self.result.created += 1

    # === RETAINAGE -> retainage_entries (36 rows, 2 sections) ===
    # Section 1 "HELD BY LENDERS" row 4: Project[0] Lender[1] Total Contract[2]
    #   Total Billed[3] Retainage%[4] Retainage Held($)[5] ...
    # Section 2 "YOU OWE TO SUBS" row 14: Sub/Vendor[0] Project[1] ...
    def _import_retainage(self, rows):
        if len(rows) < 6: return
        section = None
        for row in rows:
            val0 = self.clean_string(self._val(row, 0))
            if not val0: continue
            if "HELD BY LENDERS" in val0: section = "receivable"; continue
            if "YOU OWE" in val0 or "RETAINAGE YOU OWE" in val0: section = "payable"; continue
            if val0 in ("Project", "Sub/Vendor") or val0.startswith("TOTAL") or val0.startswith("="): continue
            if section == "receivable":
                p = self.session.query(Project).filter(Project.code == val0).first()
                if not p: continue
                pct = self._val(row, 4)
                ret_pct = Decimal(str(pct)) if pct and not isinstance(pct, str) else Decimal("0.10")
                billed = self.clean_currency(self._val(row, 3))
                held = billed * ret_pct if billed > 0 else Decimal("0")
                entry = RetainageEntry(project_id=p.id, amount_held=held, balance=held,
                    notes=f"Lender: {self._val(row, 1) or ''} | Release: {self._val(row, 6) or ''} | Type: receivable")
                self.session.add(entry); self.result.created += 1
            elif section == "payable":
                vid = self.get_or_create_vendor(val0)
                pc = self.clean_string(self._val(row, 1))
                p = self.session.query(Project).filter(Project.code == pc).first() if pc else None
                pct = self._val(row, 4)
                ret_pct = Decimal(str(pct)) if pct and not isinstance(pct, str) else Decimal("0.10")
                paid = self.clean_currency(self._val(row, 3))
                held = paid * ret_pct if paid > 0 else Decimal("0")
                entry = RetainageEntry(project_id=p.id if p else None, vendor_id=vid,
                    amount_held=held, balance=held,
                    notes=f"Release: {self._val(row, 6) or ''} | Type: payable")
                self.session.add(entry); self.result.created += 1

    # === BID PIPELINE -> bid_pipeline (46 rows) ===
    # Row 3 header, data shifted: Name[0] Client[1] Salesperson[2]
    #   Date?[3] Date2?[4] Value[5] Status[6] ...
    def _import_bid_pipeline(self, rows):
        if len(rows) < 5: return
        for row in rows[4:]:
            name = self.clean_string(self._val(row, 0))
            if not name or name.startswith("TOTAL"): self.result.skipped += 1; continue
            value = self.clean_currency(self._val(row, 5))
            status_raw = (self.clean_string(self._val(row, 6)) or "").lower()
            status = BidStatus.identified
            if "won" in status_raw: status = BidStatus.won
            elif "lost" in status_raw: status = BidStatus.lost
            elif "submitted" in status_raw: status = BidStatus.bid_submitted
            elif "pursuing" in status_raw: status = BidStatus.pursuing
            prob_raw = self._val(row, 8)
            prob = None
            if prob_raw is not None and not isinstance(prob_raw, str):
                try: prob = Decimal(str(round(float(prob_raw)*100, 2)))
                except: pass
            bid = BidPipeline(opportunity_name=name,
                client_name=self.clean_string(self._val(row, 1)),
                salesperson=self.clean_string(self._val(row, 2)),
                estimated_value=value, status=status, probability=prob,
                notes=self.clean_string(self._val(row, 16)))
            self.session.add(bid); self.result.created += 1

    # === CREW ALLOCATION -> crew_allocations (36 rows) ===
    # Row 3 header: Employee/Resource[0] Role[1] week cols [2]-[14]
    def _import_crew_allocation(self, rows):
        if len(rows) < 5: return
        date_row = rows[3] if len(rows) > 3 else None
        week_dates = []
        if date_row:
            for i in range(2, min(15, len(date_row))):
                raw = self.clean_string(self._val(date_row, i))
                d = self.clean_date(f"{raw}/2026") if raw else None
                week_dates.append((i, d))
        for row in rows[4:]:
            emp_name = self.clean_string(self._val(row, 0))
            if not emp_name or emp_name.startswith("TOTAL"): continue
            role = self.clean_string(self._val(row, 1))
            parts = emp_name.split(None, 1)
            emp = self.session.query(Employee).filter(Employee.first_name == parts[0]).first()
            eid = emp.id if emp else None
            for ci, wd in week_dates:
                assignment = self.clean_string(self._val(row, ci))
                if not assignment or wd is None: continue
                p = self.session.query(Project).filter(Project.code == assignment).first()
                pid = p.id if p else self.get_or_create_project(assignment, assignment)
                alloc = CrewAllocation(employee_id=eid, project_id=pid,
                    week_starting=wd, role_on_project=role,
                    hours_allocated=Decimal("40"), notes=emp_name)
                self.session.add(alloc); self.result.created += 1

    # === AR AGING -> invoices (28 rows) ===
    # Row 1 header: Client[0] Amount Owed[1] Invoice Date[2] Due Date[3]
    #   Description[4] Aging Bucket[5] Status[6] Collection Notes[7] SOURCE[8]
    def _import_ar_aging(self, rows):
        if len(rows) < 3: return
        for row in rows[2:]:
            client = self.clean_string(self._val(row, 0))
            if not client or client.startswith("TOTAL"): self.result.skipped += 1; continue
            amount = self.clean_currency(self._val(row, 1))
            if amount == 0: self.result.skipped += 1; continue
            bucket = self.clean_string(self._val(row, 5))
            status = InvoiceStatus.overdue
            if bucket == "0-30": status = InvoiceStatus.sent
            inv = Invoice(project_id=None,
                invoice_number=self.clean_string(self._val(row, 4)) or f"AR-{client[:10]}",
                date_issued=self.clean_date(self._val(row, 2)),
                date_due=self.clean_date(self._val(row, 3)),
                amount=amount, balance=amount, status=status,
                notes=f"Client: {client} | Bucket: {bucket} | {self._val(row, 7) or ''}")
            self.session.add(inv); self.result.created += 1

    # === AP AGING -> vendor updates (22 rows) ===
    # Row 1 header: Vendor[0] Amount Owed[1] Aging Bucket[2] Priority[3]
    #   Notes[4] Risk Level[5] Action[6] SOURCE[7]
    def _import_ap_aging(self, rows):
        if len(rows) < 3: return
        for row in rows[2:]:
            vn = self.clean_string(self._val(row, 0))
            if not vn or vn.startswith("TOTAL"): self.result.skipped += 1; continue
            amount = self.clean_currency(self._val(row, 1))
            bucket = self.clean_string(self._val(row, 2))
            priority = self.clean_string(self._val(row, 3))
            risk = self.clean_string(self._val(row, 5))
            vendor = self.session.query(Vendor).filter(Vendor.name == vn).first()
            if not vendor:
                vendor = Vendor(name=vn); self.session.add(vendor); self.session.flush()
            vendor.notes = (vendor.notes or "") + f" | AP: ${amount:,.0f} ({bucket}) Pri: {priority} Risk: {risk}"
            self.result.updated += 1

    # === LOWES PRO -> cost_events (446 rows) ===
    # Row 1 header: Date[0] Store[1] PO(Raw)[2] Project(Normalized)[3]
    #   Division[4] Purchaser[5] Invoice#[6] CC Last 4[7] Tax[8]
    #   Order Total[9] MF?[10] SOURCE[11]
    def _import_lowes_pro(self, rows):
        if len(rows) < 3: return
        lowes_id = self.get_or_create_vendor("Lowe's Pro")
        for row in rows[2:]:
            date = self.clean_date(self._val(row, 0))
            total = self.clean_currency(self._val(row, 9))
            if not date or total == 0: self.result.skipped += 1; continue
            pc = self.clean_string(self._val(row, 3))
            pid = None
            if pc:
                p = self.session.query(Project).filter(Project.code == pc).first()
                pid = p.id if p else self.get_or_create_project(pc, pc)
            event = CostEvent(project_id=pid, vendor_id=lowes_id, date=date,
                amount=-abs(total), event_type=CostEventType.material_purchase,
                description=f"Lowe's {self._val(row, 1) or ''} — PO: {self._val(row, 2) or ''}",
                reference_number=self.clean_string(self._val(row, 6)),
                source=CostEventSource.lowes_import, source_ref="Lowe's Pro Export",
                import_batch=self.batch_id,
                notes=f"Purchaser: {self._val(row, 5) or ''} | CC: {self._val(row, 7) or ''} | Tax: {self._val(row, 8) or 0} | MF: {self._val(row, 10) or 'N'}")
            self.session.add(event); self.result.created += 1
            if self.result.created % 500 == 0: self.session.flush()

    # === VICTORY CROSSINGS -> bid_pipeline (134 rows, pro forma) ===
    def _import_victory_crossings(self, rows):
        if len(rows) < 5: return
        bid = BidPipeline(opportunity_name="Victory Crossings 64-Unit Development",
            client_name="SECG (Internal Development)", project_type="Multifamily",
            estimated_value=Decimal("9256000"), status=BidStatus.pursuing,
            salesperson="Samuel Carson",
            notes="64-unit development pro forma — imported from VICTORY CROSSINGS tab")
        self.session.add(bid); self.result.created += 1
