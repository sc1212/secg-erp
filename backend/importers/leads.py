"""Importer for BuilderTrend leads and proposal exports.

Reads: Leads__1_.xlsx, LeadProposals__9_.xlsx
Creates: Lead, LeadProposal records
"""

import os
from typing import Optional

import openpyxl
from sqlalchemy.orm import Session

from backend.importers.base import BaseImporter, ImportResult
from backend.models.extended import Lead, LeadProposal


class LeadsImporter(BaseImporter):
    """Imports the Leads export from BuilderTrend."""

    source_name = "leads"
    source_type = "buildertrend_import"

    def __init__(self, session: Session, file_path: str,
                 batch_id: Optional[str] = None):
        super().__init__(session, batch_id)
        self.file_path = file_path

    def run(self) -> ImportResult:
        wb = openpyxl.load_workbook(self.file_path, read_only=True)
        ws = wb["Leads"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 3:
            self.result.errors.append("Leads file has fewer than 3 rows")
            self.result.finish()
            return self.result

        # Row 0 = title row ("ALL LEADS / Not the same as the job list")
        # Row 1 = header row
        header = [str(h).strip() if h else "" for h in rows[1]]
        col_map = {name: idx for idx, name in enumerate(header)}

        for row_data in rows[2:]:
            try:
                self._import_lead(row_data, col_map)
            except Exception as e:
                title = row_data[col_map.get("Opportunity Title", 2)] if len(row_data) > 2 else "?"
                self.result.errors.append(f"Row '{title}': {e}")

        self.session.commit()
        self.log_data_source(self.result.total_processed)
        self.session.commit()
        self.result.finish()
        return self.result

    def _import_lead(self, row: tuple, col: dict):
        """Process a single lead row."""
        title = self.clean_string(row[col.get("Opportunity Title", 2)])
        if not title:
            self.result.skipped += 1
            return

        # Dedup by opportunity title + client contact
        contact = self.clean_string(row[col.get("Client Contact", 4)])
        existing = self.session.query(Lead).filter(
            Lead.opportunity_title == title,
            Lead.client_contact == contact,
        ).first()

        if existing:
            self.result.skipped += 1
            return

        # Parse estimated revenue (handles "$0.00" format)
        est_rev_str = self.clean_string(row[col.get("Estimated Revenue", 22)])
        est_rev = self.clean_currency(est_rev_str)

        # Map lead status
        raw_status = self.clean_string(row[col.get("Lead Status", 5)])

        lead = Lead(
            opportunity_title=title,
            client_contact=contact,
            email=self.clean_string(row[col.get("Email", 1)]),
            phone=self.clean_string(row[col.get("Phone", 33)]),
            cell_phone=self.clean_string(row[col.get("Cell Phone", 16)]),
            street_address=self.clean_string(row[col.get("Street Address (Contact)", 19)]),
            city=self.clean_string(row[col.get("City (Contact)", 17)]),
            state=self.clean_string(row[col.get("State (Contact)", 18)]),
            zip_code=self.clean_string(row[col.get("Zip (Contact)", 20)]),
            opp_street_address=self.clean_string(row[col.get("Street Address(Opp)", 31)]),
            opp_city=self.clean_string(row[col.get("City(Opp)", 29)]),
            opp_state=self.clean_string(row[col.get("State(Opp)", 30)]),
            opp_zip=self.clean_string(row[col.get("Zip(Opp)", 32)]),
            lead_status=raw_status,
            confidence=self.clean_int(row[col.get("Confidence", 7)]),
            estimated_revenue_min=self.clean_currency(row[col.get("Estimated Revenue Min", 8)]),
            estimated_revenue_max=self.clean_currency(row[col.get("Estimated Revenue Max", 9)]),
            estimated_revenue=est_rev,
            salesperson=self.clean_string(row[col.get("Salesperson", 11)]),
            source=self.clean_string(row[col.get("Source", 12)]),
            project_type=self.clean_string(row[col.get("Project Type", 14)]),
            proposal_status=self.clean_string(row[col.get("Proposal Status...", 13)]),
            last_contacted=self.clean_string(row[col.get("Last Contacted", 10)]),
            has_been_contacted=self.clean_bool(row[col.get("Has Opportunity Been Contacted?", 0)]),
            created_date=self.clean_date(row[col.get("Created Date", 3)]),
            sold_date=self.clean_date(row[col.get("Sold Date", 37)]),
            projected_sales_date=self.clean_date(row[col.get("Projected Sales Date", 34)]),
            related_job=self.clean_string(row[col.get("Related Job", 36)]),
            notes=self.clean_string(row[col.get("Notes", 28)], 2000),
        )
        self.session.add(lead)
        self.result.created += 1


class ProposalsImporter(BaseImporter):
    """Imports the LeadProposals export from BuilderTrend."""

    source_name = "lead_proposals"
    source_type = "buildertrend_import"

    def __init__(self, session: Session, file_path: str,
                 batch_id: Optional[str] = None):
        super().__init__(session, batch_id)
        self.file_path = file_path

    def run(self) -> ImportResult:
        wb = openpyxl.load_workbook(self.file_path, read_only=True)
        ws = wb["Lead Proposals"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 3:
            self.result.errors.append("Proposals file has fewer than 3 rows")
            self.result.finish()
            return self.result

        # Row 0 = title ("Lead Proposals (exported on ...)")
        # Row 1 = header
        header = [str(h).strip() if h else "" for h in rows[1]]
        col_map = {name: idx for idx, name in enumerate(header)}

        for row_data in rows[2:]:
            try:
                self._import_proposal(row_data, col_map)
            except Exception as e:
                title = row_data[0] if row_data else "?"
                self.result.errors.append(f"Row '{title}': {e}")

        self.session.commit()
        self.log_data_source(self.result.total_processed)
        self.session.commit()
        self.result.finish()
        return self.result

    def _import_proposal(self, row: tuple, col: dict):
        """Process a single proposal row."""
        title = self.clean_string(row[col.get("Proposal Title", 0)])
        if not title:
            self.result.skipped += 1
            return

        # Dedup
        existing = self.session.query(LeadProposal).filter(
            LeadProposal.proposal_title == title,
        ).first()

        if existing:
            self.result.skipped += 1
            return

        # Try to link to existing lead
        opp_title = self.clean_string(row[col.get("Opportunity Title", 1)])
        lead_id = None
        if opp_title:
            lead = self.session.query(Lead).filter(
                Lead.opportunity_title == opp_title
            ).first()
            if lead:
                lead_id = lead.id

        # Parse price (handles "$211,178.16" format)
        price_str = self.clean_string(row[col.get("Client Price", 4)])
        price = self.clean_currency(price_str)

        proposal = LeadProposal(
            lead_id=lead_id,
            proposal_title=title,
            opportunity_title=opp_title,
            client_contact=self.clean_string(row[col.get("Client Contact", 2)]),
            salesperson=self.clean_string(row[col.get("Salesperson...", 3)]),
            client_price=price,
            status=self.clean_string(row[col.get("Proposal Status", 5)]),
        )
        self.session.add(proposal)
        self.result.created += 1
