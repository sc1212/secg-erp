"""Importer for Open Jobs, Scopes & Quotes from BuilderTrend export.

Reads: Open_Jobs_Next_Steps_Quotes.xlsx
Creates/updates: Project stubs (Jobs Master), Quote records (Scopes & Quotes)
"""

from typing import Optional

import openpyxl
from sqlalchemy.orm import Session

from backend.importers.base import BaseImporter, ImportResult
from backend.models.core import Project, ProjectStatus, Quote


# Maps short job codes from the Jobs Master sheet to our internal project codes
JOB_CODE_MAP = {
    "KA1": "KA1",
    "KA2": "KA2",
    "CS1": "CS1",
    "CS2": "CS2",
    "WG1": "WG1",
    "WG2": "WG2",
    "WG3": "WG3",
    "RV1": "RV1",
    "VET1": "VET1",
}


class OpenJobsImporter(BaseImporter):
    """Imports jobs and scope quotes from the Open Jobs workbook."""

    source_name = "open_jobs_quotes"
    source_type = "buildertrend_import"

    def __init__(self, session: Session, file_path: str,
                 batch_id: Optional[str] = None):
        super().__init__(session, batch_id)
        self.file_path = file_path

    def run(self) -> ImportResult:
        wb = openpyxl.load_workbook(self.file_path, read_only=True)

        # ── Import Jobs Master ───────────────────────────────────────────
        if "Jobs Master" in wb.sheetnames:
            self._import_jobs_master(wb["Jobs Master"])

        # ── Import Scopes & Quotes ───────────────────────────────────────
        if "Scopes & Quotes" in wb.sheetnames:
            self._import_quotes(wb["Scopes & Quotes"])

        wb.close()
        self.session.commit()
        self.log_data_source(self.result.total_processed)
        self.session.commit()
        self.result.finish()
        return self.result

    def _import_jobs_master(self, ws):
        """Read the Jobs Master sheet and ensure projects exist."""
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return

        # Row 0 = header: [Job ID, Job Name]
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            job_id = self.clean_string(row[0], 20)
            job_name = self.clean_string(row[1])
            if not job_id:
                continue

            # Map the job ID (e.g., "J001") to a project code
            # The job_name itself contains the short code like "KA1"
            project_code = job_name if job_name else job_id

            existing = self.session.query(Project).filter(
                Project.code == project_code
            ).first()

            if not existing:
                project = Project(
                    code=project_code,
                    name=project_code,
                    status=ProjectStatus.active,
                )
                self.session.add(project)
                self.result.created += 1
            else:
                self.result.skipped += 1

        self.session.flush()

    def _import_quotes(self, ws):
        """Read Scopes & Quotes and create Quote records."""
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return

        # Row 0 = header
        header = [str(h).strip() if h else "" for h in rows[0]]
        col = {name: idx for idx, name in enumerate(header)}

        for row in rows[1:]:
            if not row or not row[0]:
                continue
            try:
                self._import_single_quote(row, col)
            except Exception as e:
                qid = row[0] if row else "?"
                self.result.errors.append(f"Quote '{qid}': {e}")

    def _import_single_quote(self, row: tuple, col: dict):
        """Process one quote row."""
        quote_number = self.clean_string(row[col.get("Quote ID", 0)], 50)
        job_name = self.clean_string(row[col.get("Job Name", 2)])

        if not quote_number:
            self.result.skipped += 1
            return

        # Resolve the project
        project_code = job_name
        project = self.session.query(Project).filter(
            Project.code == project_code
        ).first()

        if not project:
            # Create a stub project
            project = Project(
                code=project_code or f"UNKNOWN_{quote_number}",
                name=project_code or "Unknown Project",
                status=ProjectStatus.active,
            )
            self.session.add(project)
            self.session.flush()

        # Dedup by quote number + project
        existing = self.session.query(Quote).filter(
            Quote.quote_number == quote_number,
            Quote.project_id == project.id,
        ).first()

        if existing:
            # Update amount if changed
            new_amount = self.clean_currency(row[col.get("Quote Amount", 9)])
            if new_amount != existing.amount:
                existing.amount = new_amount
                self.result.updated += 1
            else:
                self.result.skipped += 1
            return

        vendor_name = self.clean_string(row[col.get("Vendor", 7)])
        vendor_id = self.get_or_create_vendor(vendor_name) if vendor_name else None

        quote = Quote(
            project_id=project.id,
            vendor_id=vendor_id,
            quote_number=quote_number,
            scope_category=self.clean_string(row[col.get("Scope Category", 4)]),
            scope_description=self.clean_string(row[col.get("Scope Description", 5)], 2000),
            labor_material=self.clean_string(row[col.get("Labor / Material / Combined", 6)]),
            amount=self.clean_currency(row[col.get("Quote Amount", 9)]),
            quote_date=self.clean_date(row[col.get("Quote Date", 8)]),
            is_approved=self.clean_bool(row[col.get("Approved? (Yes/No)", 10)]),
            contract_issued=self.clean_bool(row[col.get("Contract Issued? (Yes/No)", 11)]),
            priority=self.clean_int(row[col.get("Priority", 3)]),
            notes=self.clean_string(row[col.get("Notes", 12)], 2000),
        )
        self.session.add(quote)
        self.result.created += 1
